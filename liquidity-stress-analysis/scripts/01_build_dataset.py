"""
01_build_dataset.py
Assemble the district-year liquidity dataset from the shared raw/cleansed data.

Outputs (into liquidity-stress-analysis/output/):
  district_year_master.csv  -- one row per district per fiscal year, STATEWIDE
                               (all CAR districts), raw fields only. FY2017-2024.
  focus_peer_detail.csv     -- the 15 large audited districts FY2020-2025 with
                               audited-grade detail (cash, components, solvency,
                               audit findings) that is not available statewide.

Derived metrics + risk scoring are added in 02_compute_metrics.py.
"""
import glob
import os
import openpyxl
import pandas as pd

import common as C


# ---------------------------------------------------------------------------
# 1. CAR fund balances -> wide GF + per-fund balances per district-year
# ---------------------------------------------------------------------------
def load_car():
    df = pd.read_csv(C.CAR_CSV)
    df = df[df["fiscal_year"].between(C.FY_MIN, C.FY_MAX)].copy()
    df["district_code"] = df["district_code"].astype(int)

    gf = df[df["fund"] == "General"].copy()
    base = gf[[
        "district_code", "district_name", "fiscal_year",
        "revenues", "expenditures", "ending_balance",
        "gf_unassigned", "gf_assigned",
    ]].rename(columns={
        "revenues": "gf_revenues",
        "expenditures": "gf_expenditures",
        "ending_balance": "gf_total_fund_balance",
        "gf_unassigned": "gf_unassigned_car",
        "gf_assigned": "gf_assigned_car",
    })

    # Per-fund ending balances (one column per fund of interest).
    def fund_balance(fund_name, col):
        sub = df[df["fund"] == fund_name][["district_code", "fiscal_year", "ending_balance"]]
        return sub.rename(columns={"ending_balance": col})

    for fund_name, col in [
        ("Management", "management_fund_balance"),
        ("PPEL", "ppel_fund_balance"),
        ("Sales Tax", "save_fund_balance"),
        ("Other Capital Projects", "other_capital_projects_balance"),
        ("Debt Service", "debt_service_fund_balance"),
    ]:
        base = base.merge(fund_balance(fund_name, col),
                          on=["district_code", "fiscal_year"], how="left")

    # Total governmental fund balance = sum of governmental funds (ex Enterprise/Nutrition).
    govt = df[df["fund"].isin(C.GOVERNMENTAL_FUNDS)]
    tot = (govt.groupby(["district_code", "fiscal_year"])["ending_balance"]
           .sum().reset_index()
           .rename(columns={"ending_balance": "total_governmental_fund_balance"}))
    base = base.merge(tot, on=["district_code", "fiscal_year"], how="left")
    return base


# ---------------------------------------------------------------------------
# 2. SBRC Final Cash Reserve Levy workbooks -> statewide assigned+unassigned,
#    cash reserve levy, 20% cap, remaining levy capacity.
#    The file labelled "FY (Y+2)" is computed from ACTUAL fiscal-year Y financials.
# ---------------------------------------------------------------------------
def load_sbrc():
    rows = []
    for path in sorted(glob.glob(str(C.SBRC_DIR / "Final Cash Reserve Levies, FY *.xlsx"))):
        file_fy = int(os.path.basename(path).split("FY ")[1].split(".")[0])
        actual_fy = file_fy - 2
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["CashReserveLevy"] if "CashReserveLevy" in wb.sheetnames else wb[wb.sheetnames[0]]
        data = list(ws.iter_rows(values_only=True))
        # header is the row containing 'FiscalYear'
        hidx = next(i for i, r in enumerate(data) if r and r[0] == "FiscalYear")
        for r in data[hidx + 1:]:
            if not r or r[2] in (None, "", "9999"):
                continue
            try:
                code = int(r[2])
            except (TypeError, ValueError):
                continue
            rows.append({
                "district_code": code,
                "fiscal_year": actual_fy,
                "cash_reserve_levy_budget_fy": file_fy,
                "sbrc_expenditures": r[7],
                "sbrc_20pct_cap": r[8],
                "gf_assigned_plus_unassigned_sbrc": r[9],
                "max_cash_reserve_levy_capacity": r[10],   # Final Maximum Cash Reserve Levy
                "cash_reserve_levy_amount": r[11],         # Final Cash Reserve Levy (levied)
                "levying_maximum_cash_reserve": r[12],
            })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# 3. UAB workbook 'CE & Chart Data' -> statewide certified enrollment + UAB.
# ---------------------------------------------------------------------------
def load_uab_enrollment():
    wb = openpyxl.load_workbook(C.UAB_XLSX, read_only=True, data_only=True)
    ws = wb["CE & Chart Data"]
    data = list(ws.iter_rows(values_only=True))
    hidx = next(i for i, r in enumerate(data) if r and r[0] == "FiscalYear")
    hdr = data[hidx]
    idx = {h: j for j, h in enumerate(hdr) if h}
    rows = []
    for r in data[hidx + 1:]:
        if not r or r[idx["Dist"]] in (None, "", "9999"):
            continue
        try:
            code = int(r[idx["Dist"]])
            fy = int(r[idx["FiscalYear"]])
        except (TypeError, ValueError):
            continue
        rows.append({
            "district_code": code,
            "fiscal_year": fy,
            "certified_enrollment": r[idx["x249"]],
            "unspent_authorized_budget": r[idx["Unspent Authorized Budget"]],
            "uab_per_pupil": r[idx["UAB Per Pupil"]],
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# 4. Audited 15-district detail (components split, cash, solvency, findings).
# ---------------------------------------------------------------------------
def load_audited():
    a = pd.read_csv(C.AUDITED_CSV)
    a["district_code"] = a["district"].map(C.AUDITED_NAME_TO_CODE)
    keep = [
        "district_code", "district", "fiscal_year",
        "gf_revenue", "gf_expenditure", "gf_total_fund_balance",
        "gf_unassigned", "gf_assigned", "cash_and_investments",
        "solvency_ratio_pct", "operating_margin_pct",
        "salaries_benefits", "salary_benefit_pct_gf",
        "go_debt_outstanding", "save_rev_bonds",
        "findings_count", "material_weakness", "significant_deficiency",
        "repeat_finding", "single_audit_findings", "gfoa_cert",
        "auditor", "report_date", "opinion_type", "confidence", "notes",
    ]
    a = a[[c for c in keep if c in a.columns]].copy()
    return a


def main():
    car = load_car()
    sbrc = load_sbrc()
    uab = load_uab_enrollment()

    master = car.merge(sbrc, on=["district_code", "fiscal_year"], how="left")
    master = master.merge(uab, on=["district_code", "fiscal_year"], how="left")

    # Canonical assigned+unassigned numerator (the "practical" numerator):
    #   prefer CAR split (assigned+unassigned, FY23-24); else SBRC combined (FY20-24).
    car_au = master["gf_assigned_car"].fillna(0) + master["gf_unassigned_car"].fillna(0)
    have_car_split = master["gf_unassigned_car"].notna()
    master["gf_assigned_plus_unassigned"] = car_au.where(
        have_car_split, master["gf_assigned_plus_unassigned_sbrc"])
    master["assigned_unassigned_source"] = have_car_split.map(
        {True: "CAR (split)", False: "SBRC (combined)"})
    master.loc[master["gf_assigned_plus_unassigned"].isna(), "assigned_unassigned_source"] = "missing"

    # Unassigned-only (statewide available FY23-24 only).
    master["gf_unassigned"] = master["gf_unassigned_car"]
    master["gf_assigned"] = master["gf_assigned_car"]

    # Focus flags.
    master["is_focus15"] = master["district_code"].isin(C.FOCUS15.keys())
    master["is_named_focus"] = master["district_code"].isin(C.NAMED_FOCUS)

    # Standardise district display name for the 15 (use canonical names).
    master["district_name_display"] = master.apply(
        lambda r: C.FOCUS15.get(r["district_code"], r["district_name"]), axis=1)

    master = master.sort_values(["district_name_display", "fiscal_year"])
    master.to_csv(C.MASTER_CSV, index=False)
    print(f"wrote {C.MASTER_CSV}  ({len(master)} rows, "
          f"{master.district_code.nunique()} districts, "
          f"FY{master.fiscal_year.min()}-{master.fiscal_year.max()})")

    # Focus-peer detail (audited grade) ------------------------------------
    aud = load_audited()
    # bring statewide DOM/CAR context columns alongside audited detail
    ctx = master[[
        "district_code", "fiscal_year", "district_name_display",
        "gf_assigned_plus_unassigned_sbrc", "cash_reserve_levy_amount",
        "max_cash_reserve_levy_capacity", "sbrc_20pct_cap",
        "unspent_authorized_budget", "certified_enrollment",
        "management_fund_balance", "ppel_fund_balance", "save_fund_balance",
        "total_governmental_fund_balance",
    ]]
    focus = aud.merge(ctx, on=["district_code", "fiscal_year"], how="left")
    focus = focus.sort_values(["district", "fiscal_year"])
    focus.to_csv(C.FOCUS_CSV, index=False)
    print(f"wrote {C.FOCUS_CSV}  ({len(focus)} rows, "
          f"{focus.district_code.nunique()} districts, "
          f"FY{focus.fiscal_year.min()}-{focus.fiscal_year.max()})")


if __name__ == "__main__":
    main()
