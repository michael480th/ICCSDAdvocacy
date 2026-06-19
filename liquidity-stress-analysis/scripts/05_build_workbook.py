"""
05_build_workbook.py  -- Deliverable 3: peer benchmark Excel workbook.
Tabs: README, Data dictionary, Raw data (statewide), Derived metrics (statewide),
      District comparison, 5-yr trend, Numerator sensitivity, Risk scoring,
      ICCSD detail, Cedar Rapids detail, Charts.
"""
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

import common as C

HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13, color="1F4E79")
RISK_FILL = {
    "Very high risk": "F4CCCC", "High risk": "FCE5CD",
    "Moderate risk": "FFF2CC", "Low risk": "D9EAD3",
}


def sheet_from_df(wb, name, df, *, title=None, freeze="A2", risk_col=None, pct_cols=()):
    ws = wb.create_sheet(name[:31])
    start = 1
    if title:
        ws.cell(1, 1, title).font = TITLE_FONT
        start = 3
    for j, col in enumerate(df.columns, 1):
        c = ws.cell(start, j, col)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for i, (_, row) in enumerate(df.iterrows(), start + 1):
        for j, col in enumerate(df.columns, 1):
            v = row[col]
            if pd.isna(v):
                v = None
            elif hasattr(v, "item"):
                v = v.item()
            ws.cell(i, j, v)
        if risk_col and risk_col in df.columns:
            rc = row[risk_col]
            fill = RISK_FILL.get(rc)
            if fill:
                for j in range(1, len(df.columns) + 1):
                    ws.cell(i, j).fill = PatternFill("solid", fgColor=fill)
    # widths
    for j, col in enumerate(df.columns, 1):
        lens = df[col].astype(str).str.len().clip(upper=34)
        maxlen = int(lens.max()) if lens.notna().any() else 10
        w = max(11, min(34, max(len(str(col)) + 2, maxlen + 1)))
        ws.column_dimensions[get_column_letter(j)].width = w
    if freeze:
        ws.freeze_panes = ws[f"A{start+1}"]
    return ws


def main():
    master = pd.read_csv(C.MASTER_CSV)
    focus = pd.read_csv(C.FOCUS_CSV)
    dd = pd.read_csv(C.OUTPUT_DIR / "data_dictionary.csv")
    t1 = pd.read_csv(C.TABLES_DIR / "table1_recent_screen.csv")
    t2 = pd.read_csv(C.TABLES_DIR / "table2_five_year_trend.csv")
    t3 = pd.read_csv(C.TABLES_DIR / "table3_numerator_sensitivity.csv")
    t4 = pd.read_csv(C.TABLES_DIR / "table4_red_flags.csv")
    bq = pd.read_csv(C.TABLES_DIR / "statewide_bottom_quartile_FY2024.csv")
    t5 = pd.read_csv(C.TABLES_DIR / "table5_fy2025_audited_peers.csv")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # README sheet
    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 110
    readme = [
        ("Iowa School District Liquidity Stress Benchmarking", TITLE_FONT),
        ("Separate analysis — NOT part of the public GitHub Pages site.", None),
        ("", None),
        ("Two populations:", Font(bold=True)),
        ("  • Statewide screen: ~330 districts/yr from the Certified Annual Report (CAR), the SBRC", None),
        ("    Final Cash Reserve Levy files, and the DOE/DOM Unspent Authorized Budget workbook.", None),
        ("    Practical (assigned+unassigned) cushion is available statewide FY2020–FY2024.", None),
        ("  • Focus peer detail: the 15 large audited districts, FY2020–FY2025, with audited", None),
        ("    fund-balance components, GF cash & investments, solvency, and audit findings.", None),
        ("", None),
        ("Primary metric: Practical days cushion = (GF assigned + unassigned) / GF expenditures × 365.", Font(bold=True)),
        ("Risk bands: 0–10 Very high · 10–20 High · 20–45 Moderate · 45–75 Low · 75+ Very low.", None),
        ("", None),
        ("IMPORTANT INTERPRETATION RULE:", Font(bold=True, color="B2182B")),
        ("These are APPARENT (annual-screen) liquidity-risk classifications, not confirmed intra-year", None),
        ("cash stress. Year-end fund balances cannot reveal intra-year cash lows. A district flagged", None),
        ("here screens as liquidity-constrained; confirming actual stress requires monthly cash-flow data.", None),
        ("", None),
        ("Do NOT treat capital balances (SAVE/Sales Tax, PPEL, Other Capital Projects, Debt Service)", None),
        ("or total governmental fund balance as operating liquidity.", None),
    ]
    for i, (text, font) in enumerate(readme, 1):
        c = ws.cell(i, 1, text)
        if font:
            c.font = font

    sheet_from_df(wb, "Data dictionary", dd, title="Data dictionary")
    sheet_from_df(wb, "District comparison", t1,
                  title="Table 1 — FY2024 liquidity screen (focus districts)", risk_col="Risk class")
    sheet_from_df(wb, "5-yr trend", t2, title="Table 2 — five-year trend", risk_col="risk_class")
    sheet_from_df(wb, "Numerator sensitivity", t3,
                  title="Table 3 — ICCSD & Cedar Rapids numerator sensitivity (FY2023)")
    sheet_from_df(wb, "Red flag summary", t4, title="Table 4 — red flag summary (FY2024)",
                  risk_col="Overall concern")
    sheet_from_df(wb, "FY2025 audited peers", t5,
                  title="FY2025 audited large-peer view (no statewide data for FY2025; "
                        "Iowa City audit not filed)", risk_col="Risk class")

    # Risk scoring tab (focus, key cols)
    riskcols = ["district", "fiscal_year", "practical_days_cushion", "conservative_reserve_ratio",
                "uab_cushion", "operating_result_pct", "enrollment_3yr_trend",
                "risk_flag_count", "cushion_band", "risk_class", "risk_rationale"]
    sheet_from_df(wb, "Risk scoring", focus[riskcols], title="Risk scoring — 15 focus districts",
                  risk_col="risk_class")

    # Statewide bottom quartile
    sheet_from_df(wb, "Statewide bottom quartile", bq,
                  title="Statewide bottom-quartile practical cushion, FY2024", risk_col="risk_class")

    # ICCSD & Cedar Rapids detail
    for code, tab in [(3141, "ICCSD detail"), (1053, "Cedar Rapids detail")]:
        sub = focus[focus.district_code == code]
        cols = ["fiscal_year", "gf_revenue", "gf_expenditure", "gf_total_fund_balance",
                "gf_unassigned", "gf_assigned", "cash_and_investments",
                "conservative_days_cushion", "practical_days_cushion", "gf_cash_days",
                "uab_cushion", "crl_reliance", "crl_capacity_used_pct",
                "operating_result_pct", "enrollment_3yr_trend", "risk_class"]
        cols = [c for c in cols if c in sub.columns]
        sheet_from_df(wb, tab, sub[cols], title=tab, risk_col="risk_class")

    # Raw + derived statewide
    sheet_from_df(wb, "Raw data (statewide)", master, title="Statewide district-year master")
    # Charts tab
    ws = wb.create_sheet("Charts")
    ws.cell(1, 1, "Visualizations").font = TITLE_FONT
    row = 3
    for img in ["1_scatter_cushion_vs_uab.png", "2_bar_practical_days.png",
                "3_trend_practical_days.png", "4_waterfall_iccsd_numerators.png",
                "5_heatmap_metrics.png", "6_bar_fy2025_practical_days.png",
                "7_iccsd_management_cash_projection.png"]:
        p = C.CHARTS_DIR / img
        if p.exists():
            pic = XLImage(str(p))
            pic.width = int(pic.width * 0.62)
            pic.height = int(pic.height * 0.62)
            ws.add_image(pic, f"A{row}")
            row += 26

    out = C.OUTPUT_DIR / "liquidity_benchmark_workbook.xlsx"
    wb.save(out)
    print("wrote", out)


if __name__ == "__main__":
    main()
