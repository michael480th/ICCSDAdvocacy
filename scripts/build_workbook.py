#!/usr/bin/env python3
"""
Build a clean, shareable Excel workbook of the Iowa district financial benchmark.
Reads /tmp/audit/cards.json (run build_analysis.py first) + the committed data/ CSVs.

Tabs: Overview · Scorecard · time-series for the KPIs where trend matters
(UAB, solvency, operating margin, enrollment, total debt, cash-reserve levy) ·
underlying data (audited, state, balance-sheet) · Sources & definitions.
"""
import json, csv, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

cards = json.load(open("/tmp/audit/cards.json"))            # sorted desc by composite
YEARS = [2020, 2021, 2022, 2023, 2024, 2025]
IC = "Iowa City CSD"
large = [c for c in cards if (c["enrollment"] or 0) >= 5000 and c["district"] != IC]

# ---- styles ----
HDR = Font(bold=True, color="FFFFFF", size=11)
HDRFILL = PatternFill("solid", fgColor="1F4E79")
TITLE = Font(bold=True, size=14, color="1F4E79")
SUB = Font(italic=True, size=9, color="666666")
BOLD = Font(bold=True)
ICFILL = PatternFill("solid", fgColor="DDEBF7")        # Iowa City highlight
AVGFILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center")

def score_fill(v):
    if v is None: return None
    if v >= 4: return PatternFill("solid", fgColor="C6EFCE")   # green
    if v >= 3: return PatternFill("solid", fgColor="FFEB9C")   # amber
    return PatternFill("solid", fgColor="FFC7CE")              # red

def hdr_row(ws, row, headers, widths=None):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HDR; c.fill = HDRFILL; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    if widths:
        for j, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(j)].width = w

wb = Workbook()

# ============================ Overview ============================
ws = wb.active; ws.title = "Overview"; ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 110
def line(txt, font=None, row=[2]):
    c = ws.cell(row=row[0], column=2, value=txt)
    if font: c.font = font
    c.alignment = Alignment(wrap_text=True, vertical="top"); row[0]+= 1
line("Iowa Large-District Financial Benchmark", TITLE)
line(f"15 of Iowa's largest school districts · fiscal years 2020–2025 · compiled {datetime.date(2026,6,2):%B %Y}", SUB)
line("")
line("WHAT THIS IS", BOLD)
line("A comparison of 15 of Iowa's largest school districts on financial health, quality of financial "
     "management, and how they are paying for buildings. Every figure traces to a district's audited "
     "financial report or an official Iowa state filing; nothing is estimated to fill gaps.")
line("")
line("SOURCES", BOLD)
line("• Audited Annual Comprehensive Financial Reports (ACFRs), FY2020–FY2025 — the 'Audit data' and 'Balance sheet & commitments' tabs.")
line("• Iowa Dept. of Management / Dept. of Education filings (spending authority/UAB, enrollment, levy rates, valuations, cash-reserve levy, at-risk) — the 'State data' tab. These are state-computed and UNAUDITED, but exist even where a district's audit is missing.")
line("")
line("HOW THE SCORES WORK (1 = weak, 5 = strong)", BOLD)
line("• Financial Health = 0.50 × spending authority (UAB) + 0.30 × reserves/solvency + 0.20 × 3-yr operating-margin trend.")
line("• Operational Quality = audit opinions, internal-control findings, repeat findings, audit timeliness, GFOA/ASBO recognition.")
line("• Capital Sustainability = 0.35 × Health + 0.20 × enrollment trend + 0.15 × margin + 0.20 × forward debt burden + 0.10 × GO-debt headroom.")
line("• Composite = 0.40 × Health + 0.35 × Quality + 0.25 × Capital Sustainability. 'Building vs. maintaining' is a label, not scored.")
line("")
line("PEER GROUPS (size-matched)", BOLD)
line("'Large districts' = the 12 districts with 5,000+ students. 'Best-run large districts' = the 5 highest-scoring of those. Iowa City (~14,400 students) is the 3rd-largest and is compared only to similarly large districts.")
line("")
line("IMPORTANT CAVEATS", BOLD)
line("• Iowa City CSD has audited data through FY2024 (its FY2024 audit was filed June 2026, about two years late, with five material weaknesses); its FY2025 audit is not yet filed. State figures (UAB, enrollment, levies) still exist for FY2025 and are shown; FY2025 audit-derived figures are blank.")
line("• A negative 'unrestricted net position' is normal for Iowa schools — it reflects long-term pension obligations (IPERS), not day-to-day insolvency.")
line("• Staff salaries/benefits as a share of budget are not broken out in these audits, so that common measure is not scored.")
line("• State (DOM) figures are unaudited; audited figures trace to each district's ACFR.")
line("")
line("Reproduce: scripts/extract_dom.py → scripts/build_analysis.py → scripts/build_workbook.py", SUB)

# ============================ Scorecard ============================
ws = wb.create_sheet("Scorecard")
ws.cell(row=1, column=1, value="Benchmark scorecard — ranked by composite score (1–5)").font = TITLE
ws.cell(row=2, column=1, value="Current snapshot. Score cells are shaded red (<3) / amber (3–4) / green (≥4). Iowa City highlighted.").font = SUB
cols = ["Rank","District","Enroll.","Size","Wealth","Enr. trend %/yr","Strategic posture",
        "UAB % (FY25)","Solvency % (latest)","Op. margin 3-yr %","Cash-res. levy % of cap",
        "Total debt $/student","SAVE yrs committed","Health","Op. Quality","Capital sust.","Composite","Flags"]
widths = [5,24,9,8,8,11,30,11,12,12,13,13,12,9,9,11,11,55]
hdr_row(ws, 4, cols, widths)
ws.freeze_panes = "B5"
for i, c in enumerate(cards, 1):
    r = 4 + i
    debt_pp = round(c["debt_last"]*1e6/c["enrollment"]) if (c.get("debt_last") is not None and c.get("enrollment")) else None
    vals = [i, c["district"], c["enrollment"], c["size"].replace("&gt;",">").replace("&lt;","<"), c["wealth"],
            c["enr_cagr"], c["label"], c["uab_last"], c["solv_last"], c["marg3"], c["crl_pct"],
            debt_pp, c.get("save_years"), c["health"], c["quality"], c["cap_sust"], c["composite"],
            "; ".join(c["flags"])]
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=j, value=v); cell.border = BORDER
        cell.alignment = Alignment(wrap_text=(j==18), vertical="top", horizontal=("left" if j in (2,7,18) else "center"))
    # number formats
    ws.cell(row=r, column=3).number_format = "#,##0"
    for col in (6,8,9,10,11): ws.cell(row=r, column=col).number_format = '0.0"%"'
    ws.cell(row=r, column=12).number_format = "$#,##0"
    ws.cell(row=r, column=13).number_format = '0.0" yrs"'
    for col in (14,15,16,17):
        ws.cell(row=r, column=col).number_format = "0.0"
        f = score_fill(vals[col-1]);
        if f: ws.cell(row=r, column=col).fill = f
    if c["district"] == IC:
        for j in range(1, 3): ws.cell(row=r, column=j).fill = ICFILL

# ============================ time-series helper ============================
def ts_sheet(name, title, key, fmt, note, totaldebt=False):
    ws = wb.create_sheet(name)
    ws.cell(row=1, column=1, value=title).font = TITLE
    ws.cell(row=2, column=1, value=note).font = SUB
    headers = ["District"] + [f"FY{y}" for y in YEARS]
    hdr_row(ws, 4, headers, [26] + [10]*len(YEARS))
    ws.freeze_panes = "B5"
    series_for = {}
    for i, c in enumerate(cards, 1):
        r = 4 + i
        nm = ws.cell(row=r, column=1, value=c["district"]); nm.border = BORDER; nm.font = BOLD if c["district"]==IC else Font()
        if totaldebt:
            go, sv = c["deep"]["go_debt"], c["deep"]["save_debt"]
            ser = [ (None if (go[k] is None and sv[k] is None) else (go[k] or 0)+(sv[k] or 0)) for k in range(len(YEARS)) ]
        else:
            ser = c["deep"][key]
        series_for[c["district"]] = ser
        for k, v in enumerate(ser):
            cell = ws.cell(row=r, column=2+k, value=v); cell.border = BORDER; cell.number_format = fmt; cell.alignment = CENTER
        if c["district"] == IC:
            for k in range(len(YEARS)+1): ws.cell(row=r, column=1+k).fill = ICFILL
    # average rows (large districts 5,000+, excl. Iowa City)
    r = 5 + len(cards) + 1
    lab = ws.cell(row=r, column=1, value="Large districts (5,000+) — average"); lab.font = BOLD; lab.fill = AVGFILL
    for k in range(len(YEARS)):
        vals = [series_for[c["district"]][k] for c in large if series_for[c["district"]][k] is not None]
        cell = ws.cell(row=r, column=2+k, value=(round(sum(vals)/len(vals),1) if vals else None))
        cell.number_format = fmt; cell.font = BOLD; cell.fill = AVGFILL; cell.alignment = CENTER
    return ws

ts_sheet("UAB %", "Spending authority (UAB) as % of max budget — Iowa's #1 health measure",
         "uab_pct", '0.0"%"', "Source: Iowa DOM Unspent Authorized Budget report (state, unaudited). Negative = unlawful / state review.")
ts_sheet("Solvency %", "Solvency — reserves as % of revenue (healthy range 5–15%)",
         "solvency", '0.0"%"', "Source: audited ACFRs. (Unassigned+Assigned GF balance) ÷ (GF revenue − AEA flow-through). Iowa City: audited through FY2023 only.")
ts_sheet("Operating margin %", "Operating margin — revenue minus spending, % of revenue",
         "op_margin", '0.0"%"', "Source: audited ACFRs. Above 0 = surplus; below 0 = drawing down reserves. Iowa City: audited through FY2023 only.")
ts_sheet("Enrollment", "Certified enrollment (drives state funding)",
         "enrollment", "#,##0", "Source: Iowa DOM certified/budget enrollment (state).")
ts_sheet("Total debt", "Total building debt outstanding (GO + SAVE), $",
         None, "$#,##0", "Source: audited ACFRs. GO + SAVE revenue bonds outstanding. Iowa City: audited through FY2023 only.", totaldebt=True)
ts_sheet("Cash-reserve levy %cap", "Cash-reserve levy as % of the 20% statutory cap",
         "crl_pct", '0.0"%"', "Source: Iowa DOM Final Cash Reserve Levies (state). Higher = leaning harder on this property tax for cash.")

# ============================ raw data tabs ============================
def csv_tab(name, path, title):
    ws = wb.create_sheet(name)
    ws.cell(row=1, column=1, value=title).font = TITLE
    rows = list(csv.reader(open(path)))
    hdr_row(ws, 3, rows[0])
    ws.freeze_panes = "A4"
    for i, row in enumerate(rows[1:], 4):
        for j, val in enumerate(row, 1):
            try: val = float(val) if ("." in val or val.lstrip("-").isdigit()) else val
            except (ValueError, AttributeError): pass
            ws.cell(row=i, column=j, value=val)
    for j in range(1, len(rows[0])+1):
        ws.column_dimensions[get_column_letter(j)].width = 16

csv_tab("Data - Audit (GF & debt)", "data/iowa-district-financials.csv",
        "Underlying audited data — one row per district-year (FY2020–FY2025)")
csv_tab("Data - Balance sheet+commit", "data/iowa-district-notes.csv",
        "Balance sheet & forward commitments (from audit notes)")
csv_tab("Data - Scores", "data/iowa-district-scorecards.csv",
        "Computed scores & flags (one row per district)")

# merge DOM time-series into one wide state-data tab
ws = wb.create_sheet("Data - State (DOM)")
ws.cell(row=1, column=1, value="Underlying Iowa state data (DOM/DE) — state-computed, unaudited").font = TITLE
import os
dom_files = ["unspent-authorized-budget.csv","certified-enrollment.csv","cash-reserve-levy.csv",
             "levy-rates-and-valuation.csv","at-risk.csv","aea-flowthrough.csv"]
merged = {}
fields = ["district","fiscal_year"]
for fn in dom_files:
    for r in csv.DictReader(open(f"data/dom/{fn}")):
        key = (r["district"], r["fiscal_year"]); merged.setdefault(key, {}).update(r)
        for k in r:
            if k not in fields: fields.append(k)
hdr_row(ws, 3, fields); ws.freeze_panes = "C4"
for i, key in enumerate(sorted(merged), 4):
    for j, fld in enumerate(fields, 1):
        v = merged[key].get(fld, "")
        try: v = float(v) if (v and (("." in v) or v.lstrip("-").isdigit())) else v
        except (ValueError, AttributeError): pass
        ws.cell(row=i, column=j, value=v)
for j in range(1, len(fields)+1): ws.column_dimensions[get_column_letter(j)].width = 15

# ============================ Sources & definitions ============================
ws = wb.create_sheet("Sources & definitions")
ws.cell(row=1, column=1, value="Sources & definitions").font = TITLE
ws.column_dimensions["A"].width = 26; ws.column_dimensions["B"].width = 70; ws.column_dimensions["C"].width = 38
hdr_row(ws, 3, ["Measure","Plain-English definition","Source / direction"])
defs = [
 ("Spending authority (UAB)","Iowa caps how much a district may spend each year. UAB is the unused room carried forward — the #1 Iowa health measure. Negative is unlawful.","DOM UAB report (unaudited) · higher better"),
 ("Solvency / reserves","General-fund savings as a share of one year's revenue. 5–15% is healthy in Iowa.","Audited ACFRs · higher better (to ~15%)"),
 ("Operating margin","Revenue minus spending, as % of revenue (3-yr avg in the scorecard).","Audited ACFRs · above 0 better"),
 ("Certified enrollment","Number of students, which drives most funding.","DOM (state) · context"),
 ("Total building debt","GO (property-tax) + SAVE (sales-tax) bonds outstanding.","Audited ACFRs · context (lower per student = lighter load)"),
 ("SAVE years committed","Years of SAVE sales-tax revenue already pledged to bonds (bonds ÷ annual SAVE revenue).","Audited ACFRs · lower = more flexibility"),
 ("Cash-reserve levy % of cap","Share of the allowed cash-reserve property tax in use.","DOM (state) · lower better for taxpayers"),
 ("Health / Quality / Capital sust.","The three 1–5 pillar scores; see Overview for formulas.","Computed · higher better"),
 ("Composite","0.40·Health + 0.35·Quality + 0.25·Capital sustainability.","Computed · higher better"),
 ("Strategic posture","Whether a district is building/expanding or maintaining — a label, not a grade.","Computed · context"),
]
for i, row in enumerate(defs, 4):
    for j, v in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=v); c.alignment = Alignment(wrap_text=True, vertical="top"); c.border = BORDER

wb.save("data/iowa-district-benchmark.xlsx")
print("Wrote data/iowa-district-benchmark.xlsx with tabs:", ", ".join(s.title for s in wb.worksheets))
