#!/usr/bin/env python3
"""
Extract General Fund salaries + benefits (object detail) from the Iowa DE Certified Annual Report
(CAR) workbooks -> data/car-salaries.csv.

Iowa audited General Funds report expenditures by FUNCTION, not object, so salaries+benefits can't be
separated from most ACFRs. The CAR workbooks (CAR/*.xlsx) carry the full function×object matrix; the
"GenExpData1" sheet has a TotSal and TotBen column per district. Salaries+benefits = TotSal + TotBen.
Because the CAR is a state filing, this covers ICCSD even though its FY24 audit isn't filed.

-> data/car-salaries.csv  (district, fiscal_year, salaries_benefits, source)
"""
import openpyxl, csv, os, re

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOKS = ["CAR/2022_2023 CAR data-for website.xlsx", "CAR/2023_2024 CAR data-for website (1).xlsx"]
NUMCODE = {261:"Ankeny CSD",882:"Burlington CSD",1053:"Cedar Rapids CSD",1337:"College CSD (Prairie)",
 1611:"Davenport CSD",1737:"Des Moines Independent CSD",1863:"Dubuque CSD",3141:"Iowa City CSD",
 3231:"Johnston CSD",3715:"Linn-Mar CSD",4581:"Muscatine CSD",5250:"Pleasant Valley CSD",
 6795:"Waterloo CSD",6822:"Waukee CSD",6957:"West Des Moines CSD"}

def num(x):
    try: return float(x)
    except (TypeError, ValueError): return None

def extract(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    fy = 2000 + int(re.search(r"FY(\d{2})", str(wb["General Exp"]["A1"].value)).group(1))
    ws = wb["GenExpData1"]
    hdr = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column+1)]
    si, bi = hdr.index("TotSal"), hdr.index("TotBen")   # total salary / total benefit columns
    out = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        dn = num(r[1])
        if dn and int(dn) in NUMCODE:
            out[NUMCODE[int(dn)]] = round((num(r[si]) or 0) + (num(r[bi]) or 0))
    return fy, out

def main():
    rows = []
    for f in WORKBOOKS:
        fy, out = extract(os.path.join(ROOT, f))
        for d, v in out.items():
            rows.append([d, fy, v, "Iowa DE Certified Annual Report (CAR) General Fund object detail: TotSal+TotBen"])
    rows.sort(key=lambda r: (r[0], r[1]))
    with open(os.path.join(ROOT, "data/car-salaries.csv"), "w", newline="") as fh:
        w = csv.writer(fh); w.writerow(["district","fiscal_year","salaries_benefits","source"]); w.writerows(rows)
    print(f"wrote data/car-salaries.csv: {len(rows)} rows")

if __name__ == "__main__":
    main()
