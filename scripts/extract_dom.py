#!/usr/bin/env python3
"""
Phase 1 — extract the Iowa DOM state-data layer for the 15 benchmarked districts.

Reads the uploaded DOM workbooks (staged from the repo's UAB/, AEA/, PropertyValuation/,
PropertyTaxRateFiles/, FinalCashReserveLevies/, AtRiskFormula/, AidandLevyTaxCertification/
folders) and writes clean per-district CSVs to data/dom/. All join on the 4-digit Dist code.
These are STATE-COMPUTED, UNAUDITED figures (exist even where audits are missing).
"""
import openpyxl, csv, glob, os

SRC = "/tmp/dom2"            # staged repo folders
OUT = "data/dom"
os.makedirs(OUT, exist_ok=True)
YEARS = (2020, 2021, 2022, 2023, 2024, 2025)

CODE = {"0261":"Ankeny CSD","0882":"Burlington CSD","1053":"Cedar Rapids CSD",
"1337":"College CSD (Prairie)","1611":"Davenport CSD","1737":"Des Moines Independent CSD",
"1863":"Dubuque CSD","3141":"Iowa City CSD","3231":"Johnston CSD","3715":"Linn-Mar CSD",
"4581":"Muscatine CSD","5250":"Pleasant Valley CSD","6795":"Waterloo CSD","6822":"Waukee CSD",
"6957":"West Des Moines CSD"}

def num(x):
    try: return float(x)
    except (TypeError, ValueError): return None

def writecsv(name, header, rows):
    with open(f"{OUT}/{name}", "w", newline="") as fh:
        w = csv.writer(fh); w.writerow(header)
        for r in rows: w.writerow(r)
    print(f"  {name:38s} {len(rows):3d} rows")

UABWB = f"{SRC}/UAB/Unspent Authorized Budget Report.xlsx"
wb = openpyxl.load_workbook(UABWB, data_only=True, read_only=True)

# ---- 1. UAB (data_UAB: AL=37 max, AM=38 uab, AJ=35 expend) ----
# ---- 2. AEA flow-through (sum cols M..S = idx 12..18) ----
uab_rows, aea_rows = [], []
ws = wb["data_UAB"]
for row in ws.iter_rows(min_row=2, values_only=True):
    fy, dist = row[0], row[1]
    if dist in CODE and fy in YEARS:
        mx, uab, exp = num(row[37]), num(row[38]), num(row[35])
        pct = round(100*uab/mx, 2) if (uab is not None and mx) else ""
        uab_rows.append([CODE[dist], int(fy), mx, uab, exp, pct])
        aea = sum(num(row[i]) or 0 for i in range(12, 19))   # AEA Sp Ed..AEA Prof Dev
        aea_rows.append([CODE[dist], int(fy), round(aea)])
uab_rows.sort(key=lambda r:(r[0],r[1])); aea_rows.sort(key=lambda r:(r[0],r[1]))
writecsv("unspent-authorized-budget.csv",
         ["district","fiscal_year","max_authorized_budget","unspent_authorized_budget","expenditures","uab_pct_of_max"], uab_rows)
writecsv("aea-flowthrough.csv", ["district","fiscal_year","aea_flowthrough"], aea_rows)

# ---- 3. Certified enrollment (CE & Chart Data, x249 = col idx 7) ----
enr_rows = []
ws = wb["CE & Chart Data"]
for row in ws.iter_rows(min_row=2, values_only=True):
    fy, dist = row[0], row[2]
    if dist in CODE and fy in YEARS:
        enr_rows.append([CODE[dist], int(fy), num(row[7])])
enr_rows.sort(key=lambda r:(r[0],r[1]))
writecsv("certified-enrollment.csv", ["district","fiscal_year","certified_enrollment"], enr_rows)

# ---- 4. Cash reserve levy: FY2022-2025 from Final files (cap + levying max); FY2020-21 from history ----
crl = {}   # (name,fy) -> dict
for fn in glob.glob(f"{SRC}/FinalCashReserveLevies/*.xlsx"):
    w = openpyxl.load_workbook(fn, data_only=True, read_only=True)["CashReserveLevy"]
    hdr_seen = False
    for row in w.iter_rows(min_row=1, values_only=True):
        if row[0] == "FiscalYear": hdr_seen = True; continue
        if not hdr_seen: continue
        fy, dist = row[0], row[2]
        if dist in CODE and fy in YEARS:
            exp, cap, fmax, final = num(row[7]), num(row[8]), num(row[10]), num(row[11])
            lev = (str(row[12]).strip().lower().startswith("max")) if len(row)>12 and row[12] else False
            crl[(CODE[dist], int(fy))] = dict(levy=final, cap=cap, fmax=fmax, lev="Y" if lev else "N", exp=exp)
# FY2020-2021 levy dollars from history (no cap available)
wh = wb["CashReserveLevyHistory"]
for fy, dist, name, val in wh.iter_rows(min_row=2, values_only=True):
    if dist in CODE and fy in (2020, 2021):
        crl.setdefault((CODE[dist], int(fy)), dict(levy=num(val), cap=None, fmax=None, lev="", exp=None))
crl_rows = []
for (name, fy) in sorted(crl):
    d = crl[(name, fy)]
    pct = round(100*d["levy"]/d["cap"], 1) if (d["levy"] is not None and d["cap"]) else ""
    crl_rows.append([name, fy, d["levy"], d["cap"], d["fmax"], d["lev"], pct])
writecsv("cash-reserve-levy.csv",
         ["district","fiscal_year","cash_reserve_levy","twenty_pct_cap","final_max_cash_reserve","levying_maximum","crl_pct_of_cap"], crl_rows)

# ---- 5. Levy rates + valuation (tax-rates Data sheet) ----
lev_rows = []
wt = openpyxl.load_workbook(f"{SRC}/PropertyTaxRateFiles/School Tax Rates, Historical - FY 2002 - FY 2026.xlsx",
                            data_only=True, read_only=True)["Data"]
seen = {}
for row in wt.iter_rows(min_row=2, values_only=True):
    fy, dist = row[0], row[1]
    if dist in CODE and fy in YEARS and (dist, fy) not in seen:   # first (main) row per district
        seen[(dist, fy)] = 1
        lev_rows.append([CODE[dist], int(fy), num(row[6]), num(row[5]), num(row[7]),
                         num(row[10]), num(row[11]), num(row[13]), num(row[14]),
                         num(row[15]), num(row[17])])
lev_rows.sort(key=lambda r:(r[0],r[1]))
writecsv("levy-rates-and-valuation.csv",
         ["district","fiscal_year","gen_fund_rate_with_isl","isl_rate","management_rate",
          "voted_ppel_rate","regular_ppel_rate","debt_service_rate","grand_total_rate",
          "net_valuation_with_ge","taxable_valuation"], lev_rows)

# ---- 6. Latest 100% assessed valuation (for 5% GO-debt limit) — AY2024-FY2026 file ----
assess_rows = []
wv = openpyxl.load_workbook(f"{SRC}/PropertyValuation/School District Assessed & Taxable Valuations by Class, AY2024-FY2026.xlsx",
                            data_only=True, read_only=True)["Assessed Non-TIF_Combined"]
hdr_seen = False
for row in wv.iter_rows(min_row=1, values_only=True):
    if row and row[0] == "FiscalYear": hdr_seen = True; continue
    if not hdr_seen: continue
    fy, dist = row[0], row[2]
    if dist in CODE:
        assess_rows.append([CODE[dist], int(fy), num(row[22]), num(row[20])])  # with G&E, 100% Net
assess_rows.sort(key=lambda r:r[0])
writecsv("assessed-valuation-latest.csv",
         ["district","fiscal_year","assessed_actual_with_ge","assessed_100pct_net"], assess_rows)

# ---- 7. At-risk supplementary weighting / dollars (per-FY files) ----
ar = {}
for fn in glob.glob(f"{SRC}/AtRiskFormula/*.xlsx"):
    w = openpyxl.load_workbook(fn, data_only=True, read_only=True)["Summary"]
    for row in w.iter_rows(min_row=3, values_only=True):
        dist = row[0]
        if dist in CODE:
            # fiscal year from the title cell (row1) embedded; infer from filename instead
            fy = int(fn.split("FY ")[-1].split(".")[0])
            if fy in YEARS:
                ar[(CODE[dist], fy)] = [num(row[2]), num(row[3]), num(row[4]), num(row[6]), num(row[7])]
ar_rows = [[k[0], k[1]] + ar[k] for k in sorted(ar)]
writecsv("at-risk.csv",
         ["district","fiscal_year","atrisk_frl_weight","atrisk_enroll_weight","atrisk_subtotal",
          "district_cost_pp","atrisk_dollars_generated"], ar_rows)

# ---- 8. Aid & Levy program summary (single year): budget enrollment (L101), DCPP (L103), state cost (L105) ----
al_rows = []
wa = openpyxl.load_workbook(glob.glob(f"{SRC}/AidandLevyTaxCertification/*.xlsx")[0],
                            data_only=True, read_only=True)["Data_AidAndLevy"]
for row in wa.iter_rows(min_row=2, values_only=True):
    dist = row[2]
    if dist in CODE:
        al_rows.append([CODE[dist], num(row[4]), num(row[6]), num(row[8])])  # L101, L103, L105
al_rows.sort(key=lambda r:r[0])
writecsv("aid-levy-summary.csv",
         ["district","budget_enrollment_L101","district_cost_pp_L103","state_cost_pp_L105"], al_rows)

print("\nPhase 1 complete — state-data layer written to data/dom/")
